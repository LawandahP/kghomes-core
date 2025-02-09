from concurrent.futures import ThreadPoolExecutor
import copy
import csv
from django.http import Http404
from django.shortcuts import render
from django.contrib.auth import get_user_model
from django.utils.decorators import method_decorator
from django.utils.translation import gettext_lazy as _
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from bff.utils import UseAuthApi
from config.permissions import IsRealtor

from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework import status, generics
from rest_framework.views import APIView
from rest_framework.exceptions import ParseError
from properties.utils import imageWorker
from units.models import Units, Assignment

from utils.responseBody import ResponseBody
from utils.utils import CustomPagination, compressImage, customResponse, logger


from .models import Property
from .filters import PropertyFilter
# from files.models import Files

from .serializers import PropertyDashboardSerializer, PropertySerializer, PropertyDetailsSerializer, PropertyUpdateSerializer
from rest_framework import viewsets
from rest_framework.decorators import action

class PropertyViewSet(ResponseBody, viewsets.ModelViewSet):
    queryset = Property.objects.all()
    serializer_class = PropertySerializer
    permission_classes = [IsRealtor]
    pagination_class = CustomPagination
    filterset_class = PropertyFilter
    search_fields = ['name']

    def get_serializer_class(self):
        if action == "partial-update":
            return PropertyUpdateSerializer
        return super().get_serializer_class()
    
    def get_queryset(self):
        return Property.objects.filter(account=self.request.user["account"]["id"])

    def create(self, request, *args, **kwargs):
        try:
            data = request.data
            user = request.user
            
            serializer = self.get_serializer(data=data, context={'request': request})
            if serializer.is_valid():
                account = user["account"]["id"]
                name = data['name']
                property = serializer.save(account=account)
                
                try:
                    with ThreadPoolExecutor(max_workers=3) as executor:
                        executor.submit(imageWorker, request, property)
                    logger.info("Property images Compressed and Uploaded Successfully")
                    
                    return Response(
                        {
                            "message": f"Property {name} Registered Successfully.",
                            "payload": serializer.data
                        },
                        status=status.HTTP_201_CREATED
                    )
                    
                except Exception as e:
                    return Response({"detail": f"An error occurred: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            error = {'detail': serializer.errors}
            return Response(error, status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"detail": f"An error occurred: {e}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def retrieve(self, request, *args, **kwargs):
        id = kwargs.get('pk')
        try:
            units = Units.objects.filter(property=id)
            occupiedUnits = Units.objects.filter(property=id, status="Occupied")
            vacantUnits = Units.objects.filter(property=id, status="Vacant")

            totalUnits = len(units)
            totalVacantUnits = len(vacantUnits)
            totalOccupiedUnits = len(occupiedUnits)

            propertyData = self.get_object()
            serializer = self.get_serializer(propertyData)
            
            useAuthApi = UseAuthApi("user-details")
            ownerData = useAuthApi.fetchUserDetails(serializer.data["owner"])
            serializer_data_copy = copy.deepcopy(serializer.data)
            serializer_data_copy['owner'] = ownerData
            
        except Property.DoesNotExist:
            error = {'detail': "Property not found"}
            return Response(error, status.HTTP_404_NOT_FOUND)
        
        return customResponse(
            payload=serializer_data_copy, totalUnits=totalUnits, 
            vacantUnits=totalVacantUnits, occupiedUnits=totalOccupiedUnits,
            status=status.HTTP_200_OK
        )




class FileUploadView(APIView):    

    def post(self, request, format=None):
        file = request.FILES.get("file")

        if file is None:
            raise ParseError('No file was uploaded')

        if file.name.endswith('.csv'):
            try:
                data = []
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.reader(decoded_file)
                header = next(reader)
                for row in reader:
                    data.append(row)
            except Exception as e:
                    raise ParseError('Error parsing CSV file')

        elif file.name.endswith('.xlsx'):
            # Process XLSX file
            try:
                workbook = load_workbook(file)
                sheet = workbook.active
                data = [list(row) for row in sheet.iter_rows(values_only=True)]
                header = data[0]
                data = data[1:]
            except Exception as e:
                raise ParseError('Error parsing XLSX file')
        
        else:
            raise ParseError('Unsupported file format')

        # Create instances of YourModel using the extracted data
        instances = []
        # for row in data:
        #     related_instance = User.objects.get(id=row[2])
        #     if not all(cell is None for cell in row):
        #         instance = Property(
        #             name=row[0] if len(row) > 0 else None,
        #             address=row[1] if len(row) > 1 else None,
        #             owner=related_instance if len(row) > 1 else None,
        #         )
        #         instances.append(instance)

        # Save the instances to the database
        Property.objects.bulk_create(instances)
        return customResponse(
            message="Data Successfully Uploaded",
            status=status.HTTP_201_CREATED
        )


from django.db.models import Sum, Count
from .models import Property
from units.models import Units
from leases.models import Bills, Invoice, Lease
from django.utils import timezone



class PropertyDashboardView(APIView):
    
    def get(self, request, format=None):
        owner_id = request.GET.get('owner_id')
        property_id = request.GET.get('property_id')

        filters = {}
        filters['account'] = request.user["account"]["id"]
        if owner_id:
            filters['owner'] = owner_id
        if property_id:
            filters['id'] = property_id

        properties = Property.objects.filter(**filters) if filters else Property.objects.all()
        total_properties = properties.count()
        
        units = Units.objects.filter(property__in=properties)
        total_units = units.count()

        occupied_units = Units.objects.filter(property__in=properties, status="Occupied").count()
        vacant_units = total_units - occupied_units

        vacancy_rate = ((total_units - occupied_units) / total_units) * 100 if total_units else 0
        occupancy_rate = (occupied_units / total_units) * 100 if total_units else 0
        
        total_deposits = Bills.objects.filter(
            invoice__lease__property__in=properties,
            item="Deposit"
        ).aggregate(total_deposit=Sum('amount'))['total_deposit'] or 0
        
        total_deposits_paid = Invoice.objects.filter(
            lease__property__in=properties,
            invoiceBills__item="Deposit"
        ).aggregate(
            total_amount_paid=Sum('amount_paid')
        )['total_amount_paid'] or 0

        active_tenants = Assignment.objects.filter(property__in=properties, vacated_date__isnull=True).count()
        
        data = {
            'total_properties': total_properties,
            'units': {
                "total": total_units,
                "occupied": occupied_units,
                "vacant": vacant_units
            },
            'occupancy_rate': occupancy_rate,
            'active_tenants': active_tenants,
            'vacancy_rate': vacancy_rate,
            'total_deposits': total_deposits,
            'total_deposits_paid': total_deposits_paid
        }

        return Response(data, status=200)

